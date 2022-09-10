import pandas as pd 
import nltk
import requests
from urllib.request import urlopen
from urllib.request import Request
from bs4 import BeautifulSoup
import re
from nltk.corpus import stopwords
from nltk.tokenize import word_tokenize
from nltk.stem import WordNetLemmatizer


df = pd.read_excel('Input.xlsx', usecols=['URL'])
print("Enter the URL_ID: ")
z = int(input())
ind = z-37

URL=df['URL'][ind]
# print(URL)
raw_request = Request(URL)

raw_request.add_header('User-agent', 'Mozilla/5.0 (Macintosh; Intel Mac OS X 10_9_3) AppleWebKit/537.75.14 (KHTML, like Gecko) Version/7.0.3 Safari/7046A194A')
raw_request.add_header('Accept', 'text/html,application/xhtml+xml,application/xml;q=0.9,image/webp,*/*;q=0.8')

file = urlopen(raw_request)

soup = BeautifulSoup(file,'lxml')
title = soup.find_all('h1')
content = soup.findAll('div',{'class':'td-post-content'})
str_cells = str(content)
str_title = str(title)
cleartitle = BeautifulSoup(str_title,'lxml').get_text()
cleartext = (BeautifulSoup(str_cells,'lxml').get_text())

def text(x,y):
    outfile = open(f'{z}.txt','w',encoding='utf-8')
    title = x
    textf = y
    outfile.write(title)
    outfile.write(textf)
    outfile.close()

text(cleartitle,cleartext)

output = pd.read_excel('Output Data Structure.xlsx')
extracted_text = open(f'{z}.txt','r',encoding='utf-8').read()

lemma = WordNetLemmatizer()
stop_words = stopwords.words('english')
def text_prep(x: str) -> list:
    corp = str(x).lower() 
    corp = re.sub('[^a-zA-Z]+',' ', corp).strip() 
    tokens = word_tokenize(corp)
    words = [t for t in tokens if t not in stop_words]
    lemmatize = [lemma.lemmatize(w) for w in words]
    return lemmatize

s = text_prep(extracted_text)
final = ' '.join(s)

from nltk.sentiment.vader import SentimentIntensityAnalyzer

sent = SentimentIntensityAnalyzer()
polar = sent.polarity_scores(final)

pos_score = polar['pos']
neg_score = polar['neg']

from textblob import TextBlob

senti = TextBlob(final)
subjectivity = senti.subjectivity
polarity = senti.polarity

import readability

result = readability.getmeasures(extracted_text,lang='en')
r_word = readability.getmeasures(final,lang='en')

words = result['sentence info']['words']
complex_words = result['sentence info']['complex_words']
sentence = result['sentence info']['sentences']

avg_sen_len =result['sentence info']['words_per_sentence']
perc_of_complex_words = complex_words/words
fog_index = result['readability grades']['GunningFogIndex']
avg_words_per_sen = result['sentence info']['words_per_sentence']
complex_words = result['sentence info']['complex_words']
words_r = r_word['sentence info']['words']
syll_per_word = result['sentence info']['syll_per_word']
pronouns = r_word['word usage']['pronoun']

str = final.split()
word_count = 0
for line in final:
    words = line.split()
    word_count += len(words)
avg_word_len = word_count/len(str)

from openpyxl import workbook,load_workbook

wb = load_workbook('Output Data Structure.xlsx')
ws = wb.active
rows = z-35

ws.cell(row=rows,column=3,value=pos_score)
ws.cell(row=rows,column=4,value=neg_score)
ws.cell(row=rows,column=5,value=polarity)
ws.cell(row=rows,column=6,value=subjectivity)
ws.cell(row=rows,column=7,value=avg_sen_len)
ws.cell(row=rows,column=8,value=perc_of_complex_words)
ws.cell(row=rows,column=9,value=fog_index)
ws.cell(row=rows,column=10,value=avg_words_per_sen)
ws.cell(row=rows,column=11,value=complex_words)
ws.cell(row=rows,column=12,value=words_r)
ws.cell(row=rows,column=13,value=syll_per_word)
ws.cell(row=rows,column=14,value=pronouns)
ws.cell(row=rows,column=15,value=avg_word_len)

wb.save('Output Data Structure.xlsx')